import openpyxl as xl

wb = xl.load_workbook("transaction.xlsx")
sheet = wb["Sheet1"]

for row in range (2, sheet.max_row + 1):
    cell = sheet.cell(row, 3)
    corrected_ans = cell.value - (0.1 * cell.value)
    corrected_ans_cell = sheet.cell(row, 4)
    corrected_ans_cell.value = corrected_ans

wb.save("transaction3.xlsx")

